"""
scrape_stsci_proposals.py
=========================
从 STScI 网站抓取 JWST approved proposal 信息。

来源: https://www.stsci.edu/jwst/science-execution/approved-programs
"""

import re
import sys
import time
import logging
import argparse
import requests
import pandas as pd
from pathlib import Path
from datetime import datetime
from bs4 import BeautifulSoup, Tag

# ── 科学类别注册表（GO 项目，按 Cycle 整理）────────────────────────────────────
#
# 用于 --list-categories 展示，以及验证 --science-categories 参数。
# GTO / DDT / ERS / JOINT 页面无科学子类别，通过 --types 整体选取。

SCIENCE_CATEGORIES: dict[int, list[str]] = {
    1: [
        "Exoplanets and Exoplanet Formation",
        "Galaxies",
        "Intergalactic Medium and the Circumgalactic Medium",
        "Large Scale Structure of the Universe",
        "Solar System Astronomy",
        "Stellar Physics and Stellar Types",
        "Stellar Populations and the Interstellar Medium",
        "Supermassive Black Holes and Active Galaxies",
    ],
    2: [
        "Exoplanets and Exoplanet Formation",
        "Galaxies",
        "Intergalactic Medium and the Circumgalactic Medium",
        "Large Scale Structure of the Universe",
        "Solar System Astronomy",
        "Stellar Physics and Stellar Types",
        "Stellar Populations and the Interstellar Medium",
        "Supermassive Black Holes and Active Galaxies",
    ],
    3: [
        "Exoplanets and Exoplanet Formation",
        "Galaxies",
        "Intergalactic Medium and the Circumgalactic Medium",
        "Large Scale Structure of the Universe",
        "Solar System Astronomy",
        "Stellar Physics and Stellar Types",
        "Stellar Populations and the Interstellar Medium",
        "Supermassive Black Holes and Active Galaxies",
    ],
    4: [
        "Exoplanet Atmospheres and Habitability",
        "Exoplanet System Formation and Dynamics",
        "Gas, Dust, and the ISM",
        "High-Redshift Galaxies and the Distant Universe",
        "Nearby Galaxies to Cosmic Noon",
        "Solar System Astronomy",
        "Stars and Stellar Populations",
        "Supermassive Black Holes and Active Galaxies",
    ],
    5: [
        "Exoplanet Atmospheres and Habitability",
        "Exoplanetary System Formation and Dynamics",
        "Gas, Dust, and the ISM",
        "High-Redshift Galaxies and the Distant Universe",
        "Nearby Galaxies to Cosmic Noon",
        "Solar System Astronomy",
        "Stars and Stellar Populations",
        "Supermassive Black Holes and Active Galaxies",
    ],
}

# ── 页面 URL 配置 ──────────────────────────────────────────────────────────────

BASE_URL = "https://www.stsci.edu"

GO_PATHS: dict[int, str] = {
    c: f"/jwst/science-execution/approved-programs/general-observers/cycle-{c}-go"
    for c in range(1, 6)
}

EXTRA_PATHS: dict[str, str] = {
    "GTO":   "/jwst/science-execution/approved-programs/guaranteed-time-observations",
    "DDT":   "/jwst/science-execution/approved-programs/directors-discretionary-time",
    "ERS":   "/jwst/science-execution/approved-ers-programs",
    "JOINT": "/jwst/science-execution/approved-programs/joint-observing-programs",
}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
    ),
}

OUTPUT_DIR = Path(__file__).parent

# ── 列名标准化映射 ─────────────────────────────────────────────────────────────

COL_MAP = {
    "ID": "program_id",
    "PID": "program_id",
    "Program Title": "title",
    "Title": "title",
    "PI & Co-PIs": "pi",
    "PI": "pi",
    "Principal Investigator": "pi",
    "Instrument/Mode": "instruments",
    "Instruments": "instruments",
    "Prime/Parallel Time": "allocated_hours",
    "Allocated Hours": "allocated_hours",
    "Exclusive Access Period": "exclusive_access_months",
    "Type": "obs_type",
    "AR?": "archival_research",
    "Coordinated With": "coordinated_with",
    "Allocating TAC": "allocating_tac",
    "Cycle": "_cycle_from_table",  # 用下划线前缀标记，避免与注入的 cycle 冲突
}

FINAL_COLUMNS = [
    "program_id", "title", "pi",
    "program_type", "cycle", "sci_category", "obs_type",
    "instruments", "allocated_hours", "exclusive_access_months",
    "archival_research", "coordinated_with", "allocating_tac",
    "program_href",
]


# ── 日志 ──────────────────────────────────────────────────────────────────────

def setup_logging() -> logging.Logger:
    log_path = OUTPUT_DIR / "scrape_stsci.log"
    fmt = "%(asctime)s  %(levelname)-8s  %(message)s"
    logging.basicConfig(
        level=logging.INFO,
        format=fmt,
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8", mode="a"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return logging.getLogger(__name__)

log = setup_logging()


# ── 工具函数 ──────────────────────────────────────────────────────────────────

def clean(text: str | None) -> str:
    return " ".join(text.split()).strip() if text else ""


def fetch_page(path: str) -> BeautifulSoup | None:
    url = BASE_URL + path
    for attempt in range(1, 4):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=30)
            resp.raise_for_status()
            return BeautifulSoup(resp.text, "lxml")
        except Exception as e:
            log.warning(f"  [尝试 {attempt}/3] {url}  {e}")
            time.sleep(3 * attempt)
    log.error(f"  获取失败: {BASE_URL + path}")
    return None


def extract_cycle_from_heading(text: str) -> int | None:
    m = re.search(r"[Cc]ycle\s+(\d+)", text)
    return int(m.group(1)) if m else None


def category_matches(page_category: str, wanted: list[str] | None) -> bool:
    """
    判断页面类别是否在用户指定的列表中。
    wanted=None 表示不过滤（接受全部）。
    匹配时忽略大小写，并支持部分包含匹配。
    """
    if wanted is None:
        return True
    low = page_category.lower()
    return any(w.lower() in low or low in w.lower() for w in wanted)


# ── HTML 表格解析 ─────────────────────────────────────────────────────────────

def parse_table_rows(table: Tag, extra: dict) -> list[dict]:
    thead = table.find("thead")
    if thead:
        header_cells = thead.find_all("th")
    else:
        first_tr = table.find("tr")
        header_cells = first_tr.find_all(["th", "td"]) if first_tr else []
    headers = [clean(th.get_text()) for th in header_cells]

    tbody = table.find("tbody") or table
    rows = []
    for tr in tbody.find_all("tr"):
        cells = tr.find_all(["td", "th"])
        if not cells:
            continue
        row = {}
        for i, cell in enumerate(cells):
            key = headers[i] if i < len(headers) else f"col_{i}"
            a = cell.find("a")
            if a and a.get("href"):
                row["program_href"] = a["href"]
            row[key] = clean(cell.get_text())
        row.update(extra)
        rows.append(row)
    return rows


# ── GO 页面解析 ───────────────────────────────────────────────────────────────

def scrape_go(cycle: int, wanted_categories: list[str] | None) -> pd.DataFrame:
    path = GO_PATHS[cycle]
    log.info(f"  [GO Cycle {cycle}] {BASE_URL + path}")
    soup = fetch_page(path)
    if soup is None:
        return pd.DataFrame()

    records = []
    current_category = ""
    content = soup.find("main") or soup.find("article") or soup.body
    if content is None:
        return pd.DataFrame()

    for elem in content.find_all(["h2", "h3", "h4", "table"]):
        if elem.name in ("h2", "h3", "h4"):
            current_category = clean(elem.get_text())
        elif elem.name == "table":
            if not category_matches(current_category, wanted_categories):
                continue
            extra = {"program_type": "GO", "cycle": cycle, "sci_category": current_category}
            records.extend(parse_table_rows(elem, extra))

    df = pd.DataFrame(records)
    n_cats = df["sci_category"].nunique() if not df.empty else 0
    log.info(f"    → {len(df)} 条记录，涵盖 {n_cats} 个科学类别")
    return df


# ── GTO / DDT / ERS / JOINT 页面解析 ─────────────────────────────────────────

def scrape_extra(program_type: str, target_cycles: list[int]) -> pd.DataFrame:
    """
    GTO/DDT/ERS/JOINT 页面无科学子类别，通过 Cycle 过滤。
    两种页面格式：
    - 标题格式 (GTO/JOINT): h4 含 "Cycle N" → 按标题注入 cycle
    - 列格式 (DDT): 表格含 "Cycle" 列 → 解析后按列过滤
    """
    path = EXTRA_PATHS[program_type]
    log.info(f"  [{program_type}] {BASE_URL + path}")
    soup = fetch_page(path)
    if soup is None:
        return pd.DataFrame()

    content = soup.find("main") or soup.find("article") or soup.body
    if content is None:
        return pd.DataFrame()

    records = []
    current_cycle: int | None = None
    current_label = ""

    for elem in content.find_all(["h2", "h3", "h4", "table"]):
        if elem.name in ("h2", "h3", "h4"):
            text = clean(elem.get_text())
            c = extract_cycle_from_heading(text)
            if c is not None:
                current_cycle = c
            else:
                current_label = text
        elif elem.name == "table":
            extra = {
                "program_type": program_type,
                "cycle": current_cycle,
                "sci_category": current_label,
            }
            records.extend(parse_table_rows(elem, extra))

    if not records:
        log.info(f"    → 0 条记录")
        return pd.DataFrame()

    df = pd.DataFrame(records)

    # DDT 的 Cycle 在表格列里，覆盖注入的 None
    cycle_table_col = next(
        (c for c in df.columns if c.lower() == "cycle" and c != "cycle"), None
    )
    if cycle_table_col and df["cycle"].isna().all():
        df["cycle"] = pd.to_numeric(df[cycle_table_col], errors="coerce")

    # 按目标 Cycle 过滤
    df = df[df["cycle"].isin(target_cycles)]
    log.info(f"    → {len(df)} 条记录")
    return df


# ── 列名标准化 ────────────────────────────────────────────────────────────────

def normalize(df: pd.DataFrame) -> pd.DataFrame:
    df = df.rename(columns={k: v for k, v in COL_MAP.items() if k in df.columns})

    # 合并 rename 产生的重复列（如 "ID" 和 "PID" 都映射到 "program_id"）
    if df.columns.duplicated().any():
        deduped: dict[str, pd.Series] = {}
        for col in df.columns:
            if col not in deduped:
                same_cols = df.loc[:, df.columns == col]
                deduped[col] = same_cols.bfill(axis=1).iloc[:, 0]
        df = pd.DataFrame(deduped)

    # 丢掉从表格 Cycle 列映射的冗余列
    if "_cycle_from_table" in df.columns:
        df = df.drop(columns=["_cycle_from_table"])

    # 统一列顺序
    present = [c for c in FINAL_COLUMNS if c in df.columns]
    extra_cols = [c for c in df.columns if c not in FINAL_COLUMNS]
    df = df[present + extra_cols]

    if "program_id" in df.columns:
        df["program_id"] = pd.to_numeric(df["program_id"], errors="coerce").astype("Int64")
    if "cycle" in df.columns:
        df["cycle"] = pd.to_numeric(df["cycle"], errors="coerce").astype("Int64")

    return df


# ── 输出 ──────────────────────────────────────────────────────────────────────

def save_outputs(df: pd.DataFrame, cycles: list[int], types: list[str]):
    cycle_label = "all" if sorted(cycles) == list(range(1, 6)) else "-".join(str(c) for c in sorted(cycles))
    type_label  = "all" if len(types) == len(EXTRA_PATHS) + 1 else "-".join(sorted(types))
    stem = f"jwst_proposals_c{cycle_label}_{type_label}"

    csv_path   = OUTPUT_DIR / f"{stem}.csv"
    excel_path = OUTPUT_DIR / f"{stem}.xlsx"

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    log.info(f"已保存 CSV:   {csv_path}  ({len(df):,} 行)")

    _save_excel(df, excel_path)


def _save_excel(df: pd.DataFrame, path: Path):
    from openpyxl.styles import PatternFill, Font, Alignment

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="All")
        for pt in sorted(df.get("program_type", pd.Series(dtype=str)).dropna().unique()):
            sub = df[df["program_type"] == pt].reset_index(drop=True)
            sub.to_excel(writer, index=False, sheet_name=pt[:31])

        for ws in writer.sheets.values():
            ws.freeze_panes = "A2"
            hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            hfont = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = hfill; cell.font = hfont
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            alt = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                if i % 2 == 0:
                    for cell in row:
                        cell.fill = alt
            for col in ws.columns:
                max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 60)

    log.info(f"已保存 Excel: {path}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="uv run python scrape_stsci_proposals.py",
        description="抓取 STScI JWST approved proposals，支持按 Cycle、Programmatic Category 和科学子类别过滤。",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 抓取 Cycle 4 和 5 的所有类型、所有科学类别
  uv run python scrape_stsci_proposals.py --cycle 4 5

  # 只抓取 Cycle 5 的 GO 项目中的两个科学类别
  uv run python scrape_stsci_proposals.py --cycle 5 --types go \\
      --science-categories "Galaxies" "Supermassive Black Holes and Active Galaxies"

  # 列出所有 Cycle 的可用科学类别（不下载数据）
  uv run python scrape_stsci_proposals.py --list-categories

  # 抓取所有 Cycle 的 DDT 和 JOINT 项目
  uv run python scrape_stsci_proposals.py --all-cycles --types ddt joint
""",
    )

    cycle_group = p.add_mutually_exclusive_group(required=False)
    cycle_group.add_argument(
        "--cycle", type=int, nargs="+", metavar="N",
        help="指定 Cycle（如 --cycle 4 5），可选 1-5",
    )
    cycle_group.add_argument(
        "--all-cycles", action="store_true",
        help="抓取所有 Cycle（1-5）",
    )

    p.add_argument(
        "--types", nargs="+",
        choices=["go", "gto", "ddt", "ers", "joint"],
        default=["go", "gto", "ddt", "ers", "joint"],
        metavar="TYPE",
        help=(
            "Programmatic Category（多选），可选: go gto ddt ers joint。"
            "默认全选（不含 calibration）。"
        ),
    )
    p.add_argument(
        "--science-categories", nargs="+", metavar="CATEGORY",
        help=(
            "GO 项目的科学子类别（多选，支持部分名称匹配）。"
            "不指定则抓取所有科学类别。"
            "示例: --science-categories \"Galaxies\" \"Supermassive Black Holes and Active Galaxies\""
        ),
    )
    p.add_argument(
        "--list-categories", action="store_true",
        help="列出每个 Cycle 下所有可用的科学子类别，然后退出（不抓取数据）",
    )
    return p


def cmd_list_categories():
    print("\nJWST GO 项目 — 各 Cycle 科学类别列表\n" + "=" * 50)
    for cycle, cats in SCIENCE_CATEGORIES.items():
        print(f"\n  Cycle {cycle}:")
        for cat in cats:
            print(f"    - {cat}")
    print("\nGTO / DDT / ERS / JOINT 页面无科学子类别，通过 --types 整体选取。\n")


# ── 主流程 ────────────────────────────────────────────────────────────────────

def main():
    parser = build_parser()
    args = parser.parse_args()

    # --list-categories 不需要其他参数
    if args.list_categories:
        cmd_list_categories()
        return

    # 校验：必须指定 cycle
    if not args.all_cycles and not args.cycle:
        parser.error("必须指定 --cycle N [N ...] 或 --all-cycles")

    cycles: list[int] = list(range(1, 6)) if args.all_cycles else sorted(set(args.cycle))
    types:  list[str] = [t.upper() for t in args.types]
    wanted_cats: list[str] | None = args.science_categories  # None = 不过滤

    log.info("=" * 65)
    log.info("  JWST Proposal 爬虫 — stsci.edu")
    log.info(f"  Cycles: {cycles}")
    log.info(f"  Types:  {types}")
    if wanted_cats:
        log.info(f"  科学类别过滤: {wanted_cats}")
    else:
        log.info("  科学类别过滤: 无（抓取全部）")
    log.info(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("=" * 65)

    frames: list[pd.DataFrame] = []

    # ── GO ──────────────────────────────────────────────────────────────────
    if "GO" in types:
        log.info("\n>>> GO Programs")
        for c in cycles:
            df = scrape_go(c, wanted_cats)
            if not df.empty:
                frames.append(df)
            time.sleep(1.0)

    # ── GTO / DDT / ERS / JOINT ─────────────────────────────────────────────
    for pt in ["GTO", "DDT", "ERS", "JOINT"]:
        if pt not in types:
            continue
        log.info(f"\n>>> {pt} Programs")
        df = scrape_extra(pt, cycles)
        if not df.empty:
            frames.append(df)
        time.sleep(1.0)

    if not frames:
        log.error("未获取到任何数据，请检查参数后重试。")
        sys.exit(1)

    # ── 合并整理 ─────────────────────────────────────────────────────────────
    combined = pd.concat(frames, ignore_index=True)
    combined = normalize(combined)
    combined = combined.dropna(how="all")
    sort_keys = [k for k in ["cycle", "program_type", "program_id"] if k in combined.columns]
    combined = combined.sort_values(sort_keys, na_position="last").reset_index(drop=True)

    # ── 保存 ─────────────────────────────────────────────────────────────────
    save_outputs(combined, cycles, types)

    # ── 摘要 ─────────────────────────────────────────────────────────────────
    log.info("\n" + "=" * 65)
    log.info(f"  总记录数:        {len(combined):,}")
    if "program_id" in combined.columns:
        log.info(f"  唯一 Program ID: {combined['program_id'].nunique():,}")
    if "program_type" in combined.columns:
        log.info("\n  各 Programmatic Category 记录数:")
        for pt, cnt in combined["program_type"].value_counts().items():
            log.info(f"    {pt:<10} {cnt:>6,}")
    if "sci_category" in combined.columns:
        log.info("\n  科学类别分布 (GO):")
        go_cats = combined[combined["program_type"] == "GO"]["sci_category"].value_counts()
        for cat, cnt in go_cats.items():
            log.info(f"    {str(cat):<55} {cnt:>5,}")
    log.info("=" * 65)
    log.info(f"  完成: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
